import requests
from lxml import etree
import re
from openpyxl import Workbook

base_url = "https://www.acc5.com/ask/question"

r = requests.get(base_url)
html = etree.HTML(r.text)
nums = html.xpath("//ul[@class='list-inline pagination-df']/li/a[@class='last']/@href")
nums = int(re.split(r"[./]", nums[0])[-2])

#get url by page
page_url_list = []
for p in range(1, nums+1):
    page_url_list.append(base_url+"/"+str(p)+".html")

global row_num
row_num = 1
wb = Workbook()
ws = wb.active
def write_to_xlsx(url_list):
    global row_num
    for url in url_list:
        r = requests.get(url)
        html = etree.HTML(r.text)
        question = ",".join(html.xpath("//h1[@class='ask-detail-title text-weight mb-25']//text()"))
        answer = html.xpath("//div[@class='col-xs-11']/p[@class='reply-text']//text()")
        #print(answer)
        if answer != []:
            ws.cell(row=row_num, column=1).value=question
            ws.cell(row=row_num, column=2).value=",".join(answer)
            row_num = row_num+1

#get url by question id
for url in page_url_list:
    print(url)
    r = requests.get(url)
    html = etree.HTML(r.text)
    q_list = html.xpath("//div[@class='porela']/a/@href")
    write_to_xlsx([x for x in q_list if "question" in x])
    wb.save("qa.xlsx")
  
